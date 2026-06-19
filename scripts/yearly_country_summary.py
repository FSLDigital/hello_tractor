"""Build yearly country performance metrics from treasury_data.xlsx."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "public" / "data" / "treasury_data.xlsx"
OUTPUT_XLSX = ROOT / "public" / "data" / "treasury_yearly_country_summary.xlsx"
OUTPUT_CSV = ROOT / "public" / "data" / "treasury_yearly_country_summary.csv"

# The workbook was inspected on this date. Months after June 2026 are future.
AS_OF_DATE = date(2026, 6, 18)
AS_OF_KEY = AS_OF_DATE.year * 100 + AS_OF_DATE.month


def tractor_key(row: pd.Series) -> str:
    """Use a stable tractor ID where available; historical Kenya falls back to name."""
    for column in ("tractor_id", "tractor_id_lu"):
        value = row[column]
        if pd.notna(value):
            return f"id:{int(value)}"
    return f"name:{str(row['name']).strip().casefold()}"


def add_future_covenants(performance: pd.DataFrame) -> pd.DataFrame:
    """Project future covenant hectares using the app's latest-known-covenant logic."""
    data = performance.copy()
    data["period_key"] = data["year"] * 100 + data["month_num"]
    data["tractor_key"] = data.apply(tractor_key, axis=1)

    historical = data[data["period_key"] <= AS_OF_KEY].sort_values("period_key")
    positive = historical[historical["monthly_covenant_target"].fillna(0) > 0]

    latest_by_tractor = (
        positive.groupby("tractor_key", as_index=True)["monthly_covenant_target"].last()
    )
    country_average = (
        positive.groupby("country", as_index=True)["monthly_covenant_target"].mean()
    )

    def projected_value(row: pd.Series) -> float:
        if row["period_key"] <= AS_OF_KEY:
            return 0.0
        tractor_value = latest_by_tractor.get(row["tractor_key"], np.nan)
        if pd.notna(tractor_value) and tractor_value > 0:
            return float(tractor_value)
        fallback = country_average.get(row["country"], 0.0)
        return float(fallback) if pd.notna(fallback) else 0.0

    data["future_covenants"] = data.apply(projected_value, axis=1)
    return data


def build_summary(performance: pd.DataFrame) -> pd.DataFrame:
    data = add_future_covenants(performance)

    expected = pd.to_numeric(data["expected_collection"], errors="coerce")
    paid = pd.to_numeric(data["total_collection"], errors="coerce").combine_first(
        pd.to_numeric(data["actual_collection"], errors="coerce")
    )
    data["amount_paid_resolved"] = paid.fillna(0)
    data["row_repayment_rate"] = np.where(expected > 0, paid / expected, np.nan)

    grouped = data.groupby(["country", "year"], as_index=False).agg(
        amount_paid=("amount_paid_resolved", "sum"),
        expected_collection=("expected_collection", "sum"),
        ha_worked=("monthly_area_serviced", "sum"),
        covenants=("monthly_covenant_target", "sum"),
        future_covenants=("future_covenants", "sum"),
        repayment_rows=("row_repayment_rate", "count"),
    )
    grouped["repayment_rate_avg"] = np.where(
        grouped["expected_collection"] > 0,
        grouped["amount_paid"] / grouped["expected_collection"],
        np.nan,
    )

    def count_tractors(group: pd.DataFrame) -> int:
        ids = group["tractor_id"].combine_first(group["tractor_id_lu"])
        id_coverage = ids.notna().mean()
        if id_coverage >= 0.80:
            return int(ids.nunique())
        return int(group["name"].astype(str).str.strip().str.casefold().nunique())

    tractor_counts = (
        data.groupby(["country", "year"])
        .apply(count_tractors, include_groups=False)
        .rename("no_of_tractors")
        .reset_index()
    )
    grouped = grouped.merge(tractor_counts, on=["country", "year"], how="left")

    # The normalized source workbook has no Booked/ha_booked field.
    grouped["booked"] = np.nan
    grouped = grouped[
        [
            "country",
            "year",
            "amount_paid",
            "ha_worked",
            "no_of_tractors",
            "repayment_rate_avg",
            "covenants",
            "booked",
            "future_covenants",
            "repayment_rows",
        ]
    ]
    return grouped.sort_values(["country", "year"]).reset_index(drop=True)


def write_outputs(summary: pd.DataFrame) -> None:
    display = summary.rename(
        columns={
            "country": "Country",
            "year": "Year",
            "amount_paid": "Amount Paid (Local Currency)",
            "ha_worked": "Ha Worked",
            "no_of_tractors": "No. of Tractors",
            "repayment_rate_avg": "Repayment Rate (Avg)",
            "covenants": "Covenants",
            "booked": "Booked",
            "future_covenants": "Future Covenants",
            "repayment_rows": "Repayment Rate Rows",
        }
    )

    notes = pd.DataFrame(
        [
            ["Source", str(SOURCE)],
            ["Reporting cutoff", AS_OF_DATE.isoformat()],
            [
                "Amount Paid",
                "SUM(COALESCE(total_collection, actual_collection, 0)). Older Kenya rows use actual_collection while newer rows/countries use total_collection. Values remain in each country's local currency.",
            ],
            ["Ha Worked", "SUM(monthly_area_serviced)."],
            [
                "No. of Tractors",
                "Distinct tractor IDs per country/year when at least 80% of rows have an ID; otherwise distinct owner names are used for that country/year. This avoids double-counting mixed ID/name history.",
            ],
            [
                "Repayment Rate (Avg)",
                "Portfolio-level rate: SUM(COALESCE(total_collection, actual_collection, 0)) / SUM(expected_collection) for each country and year. Rates are not capped at 100%, so aggregate overpayments can produce rates above 100%.",
            ],
            [
                "Repayment-rate alternative",
                "This is a ratio of sums, not an arithmetic average of row-level repayment rates.",
            ],
            ["Covenants", "SUM(monthly_covenant_target) recorded in the source workbook."],
            [
                "Booked",
                "Unavailable in treasury_data.xlsx: HT_Performance has no Booked/ha_booked column. Values are intentionally blank.",
            ],
            [
                "Future Covenants",
                "For rows after June 2026, use the tractor's latest non-zero monthly_covenant_target through June 2026; fall back to the country's average positive historical covenant. This follows the forecast logic in lib/data.ts.",
            ],
            [
                "Duplicate rows",
                "Rows are included as stored. The source flags 46 duplicate entries; excluding them would change financial and hectare totals and requires a separate business decision.",
            ],
        ],
        columns=["Metric", "Definition / comment"],
    )

    display.to_csv(OUTPUT_CSV, index=False)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        display.to_excel(writer, sheet_name="Yearly Summary", index=False)
        notes.to_excel(writer, sheet_name="Calculation Notes", index=False)

    workbook = load_workbook(OUTPUT_XLSX)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in worksheet.columns:
            width = min(
                max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2,
                80,
            )
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    summary_sheet = workbook["Yearly Summary"]
    headers = {cell.value: cell.column for cell in summary_sheet[1]}
    for row in range(2, summary_sheet.max_row + 1):
        summary_sheet.cell(row, headers["Amount Paid (Local Currency)"]).number_format = "#,##0.00"
        summary_sheet.cell(row, headers["Ha Worked"]).number_format = "#,##0.00"
        summary_sheet.cell(row, headers["Repayment Rate (Avg)"]).number_format = "0.00%"
        summary_sheet.cell(row, headers["Covenants"]).number_format = "#,##0.00"
        summary_sheet.cell(row, headers["Future Covenants"]).number_format = "#,##0.00"
        summary_sheet.cell(row, headers["Booked"]).value = "N/A"

    workbook.save(OUTPUT_XLSX)


def main() -> None:
    performance = pd.read_excel(SOURCE, sheet_name="HT_Performance")
    summary = build_summary(performance)
    write_outputs(summary)
    print(summary.to_string(index=False))
    print(f"\nWrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
